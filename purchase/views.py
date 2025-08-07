import json
import requests
import os
import io

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from urllib.parse import quote

from django.http import HttpResponse
from django.utils import timezone
from django.utils.text import slugify
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.core.validators import EmailValidator, RegexValidator
from django.core.exceptions import ValidationError as DjangoValidationError

from django_filters.rest_framework import DjangoFilterBackend
from django_tenants.utils import tenant_context
from rest_framework import viewsets, status, mixins, filters, permissions, serializers
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.decorators import action, permission_classes
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.response import Response
from drf_spectacular.utils import extend_schema, extend_schema_view

from companies.permissions import HasTenantAccess
from core.utils import enforce_tenant_schema
from inventory.models import IncomingProduct, Location, IncomingProductItem
from users.models import TenantUser
from users.module_permissions import HasModulePermission
from users.utils import convert_to_base64
from shared.viewsets.soft_delete_search_viewset import SearchDeleteViewSet, SearchViewSet
from .models import (PurchaseRequest, PurchaseRequestItem, Department, Vendor,
                     Product, RequestForQuotation, RequestForQuotationItem, UnitOfMeasure, PurchaseOrder, PurchaseOrderItem, PRODUCT_CATEGORY, Currency)
from .serializers import (PurchaseRequestSerializer, VendorSerializer, ProductSerializer,
                          RequestForQuotationSerializer, RequestForQuotationItemSerializer,
                          UnitOfMeasureSerializer, PurchaseRequestItemSerializer,
                          PurchaseOrderSerializer, PurchaseOrderItemSerializer,
                          ExcelUploadSerializer, CurrencySerializer)
from .utils import generate_model_pdf
from users.config import basic_action_permission_map


@extend_schema_view(
    list=extend_schema(tags=['Currency']),
    retrieve=extend_schema(tags=['Currency']),
    create=extend_schema(tags=['Currency']),
    update=extend_schema(tags=['Currency']),
    partial_update=extend_schema(tags=['Currency']),
    destroy=extend_schema(tags=['Currency']),
)
class CurrencyViewSet(SearchDeleteViewSet):
    serializer_class = CurrencySerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = Currency.objects.all()

@extend_schema_view(
    list=extend_schema(tags=['Unit Of Measure']),
    retrieve=extend_schema(tags=['Unit Of Measure']),
    create=extend_schema(tags=['Unit Of Measure']),
    update=extend_schema(tags=['Unit Of Measure']),
    partial_update=extend_schema(tags=['Unit Of Measure']),
    destroy=extend_schema(tags=['Unit Of Measure']),
)
class UnitOfMeasureViewSet(SearchDeleteViewSet):
    queryset = UnitOfMeasure.objects.all()
    serializer_class = UnitOfMeasureSerializer
    app_label = "purchase"
    model_name = "unitofmeasure"
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    search_fields = ['unit_name', 'unit_category']    
    action_permission_map = basic_action_permission_map


@extend_schema_view(
    list=extend_schema(tags=['Vendors']),
    retrieve=extend_schema(tags=['Vendors']),
    create=extend_schema(tags=['Vendors']),
    update=extend_schema(tags=['Vendors']),
    partial_update=extend_schema(tags=['Vendors']),
    destroy=extend_schema(tags=['Vendors']),
)
class VendorViewSet(SearchDeleteViewSet):
    queryset = Vendor.objects.all()
    serializer_class = VendorSerializer
    app_label = "purchase"
    model_name = "vendor"
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    parser_classes = (MultiPartParser, FormParser)
    search_fields = ['company_name', 'email']
    action_permission_map = {
        **basic_action_permission_map,
        "upload_excel" : "create",
        "upload_profile_picture": "create",
        "download_template": "view",
        }

    def handle_profile_picture(self, validated_data):
        if validated_data.get("profile_picture_image", None):
            validated_data["profile_picture"] = convert_to_base64(validated_data["profile_picture_image"])
            validated_data.pop("profile_picture_image")
        return validated_data

    def create(self, request, *args, **kwargs):
        serializer = VendorSerializer(data=request.data)
        if serializer.is_valid():
            validated_data = self.handle_profile_picture(serializer.validated_data)
            vendor = Vendor.objects.create(**validated_data)
            return Response({
                "message": "Vendor created successfully",
                "vendor": VendorSerializer(vendor, context={'request': request}).data
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = VendorSerializer(instance, data=request.data, partial=partial,
                                      context={'request': request, 'skip_validation': True})

        if serializer.is_valid():
            validated_data = self.handle_profile_picture(serializer.validated_data)
            for attr, value in validated_data.items():
                setattr(instance, attr, value)
            instance.save()
            return Response({
                "message": "Vendor updated successfully",
                "vendor": VendorSerializer(instance, context={'request': request}).data
            }, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['GET'], url_path='download-template')
    def download_template(self, request):
        """
        Endpoint to download a template Excel file for vendor import.
        The workbook will have:
        - An 'Instructions' sheet as the first sheet.
        - A 'Vendors' sheet for data entry (and import).
        - Data validation for email and phone number columns.
        """


        wb = Workbook()
        ws_instructions = wb.active
        ws_instructions.title = "Instructions"
        ws_instructions["A1"] = "Instructions for Filling the Vendors Sheet:"
        ws_instructions["A2"] = "1. Fill each row in the 'Vendors' sheet with vendor details."
        ws_instructions["A3"] = "2. All columns are required."
        ws_instructions["A4"] = "3. Do not modify the header row."
        ws_instructions["A5"] = "4. Email and phone number fields are validated. Invalid entries will be highlighted."
        ws_instructions["A7"] = "After filling, upload this file using the import feature in the system."

        # Add the Vendors sheet as the second sheet
        ws_vendors = wb.create_sheet(title="Vendors")
        headers = [
            "company_name",
            "email",
            "address",
            "phone_number",
        ]
        ws_vendors.append(headers)

        # Email validation (simple regex for demonstration)
        email_dv = DataValidation(
            type="custom",
            formula1='=ISNUMBER(MATCH("*@*.?*",INDIRECT("RC",FALSE),0))',
            showErrorMessage=True,
            errorTitle="Invalid Email",
            error="Please enter a valid email address."
        )
        ws_vendors.add_data_validation(email_dv)
        email_dv.add(f"B2:B1048576")

        # Phone number validation (digits only, length 7-15)
        phone_dv = DataValidation(
            type="custom",
            formula1='=AND(ISNUMBER(--INDIRECT("RC",FALSE)),LEN(INDIRECT("RC",FALSE))>=7,LEN(INDIRECT("RC",FALSE))<=15)',
            showErrorMessage=True,
            errorTitle="Invalid Phone Number",
            error="Phone number must be digits only, 7-15 characters."
        )
        ws_vendors.add_data_validation(phone_dv)
        phone_dv.add(f"D2:D1048576")

        # Save workbook to a BytesIO stream
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=vendor_import_template.xlsx'
        return response

    @action(detail=False, methods=['POST'], serializer_class=ExcelUploadSerializer)
    def upload_excel(self, request):
        serializer = ExcelUploadSerializer(data=request.data)
        if serializer.is_valid():
            excel_file = serializer.validated_data.get('file')
            if not excel_file or not isinstance(excel_file, InMemoryUploadedFile):
                return Response({"error": "No file provided or invalid file format"}, status=400)

            try:
                workbook = load_workbook(excel_file)
                if 'Vendors' not in workbook.sheetnames:
                    return Response({"error": "No 'Vendors' sheet found in the uploaded file."}, status=400)
                sheet = workbook['Vendors']

                errors = []
                rows_to_create = []
                update_instances = []
                BATCH_SIZE = 500

                email_validator = EmailValidator()
                phone_validator = RegexValidator(
                    regex=r'^\d{7,15}$',
                    message="Phone number must be digits only, 7-15 characters."
                )

                # Check if the sheet is empty (only header or no data)
                if sheet.max_row < 2:
                    return Response({"error": "The 'Vendors' sheet is empty."}, status=400)

                for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    company_name, email, address, phone_number = row[:4]
                    row_errors = []

                    # Check if the entire row is empty
                    if all(cell is None or str(cell).strip() == "" for cell in (company_name, email, address, phone_number)):
                        row_errors.append("Entire row is empty.")
                        errors.append({"row": idx, "values": row, "errors": row_errors})
                        continue

                    # Check for empty columns
                    empty_columns = []
                    if not company_name or str(company_name).strip() == "":
                        empty_columns.append("company_name")
                    if not email or str(email).strip() == "":
                        empty_columns.append("email")
                    if not address or str(address).strip() == "":
                        empty_columns.append("address")
                    if not phone_number or str(phone_number).strip() == "":
                        empty_columns.append("phone_number")
                    if empty_columns:
                        row_errors.append(f"Missing required fields: {', '.join(empty_columns)}. Row values: {row}")

                    # Email validation
                    if email and str(email).strip() != "":
                        try:
                            email_validator(email)
                        except DjangoValidationError:
                            row_errors.append(f"Invalid email: {email}")

                    # Phone number validation
                    if phone_number and str(phone_number).strip() != "":
                        try:
                            phone_validator(str(phone_number))
                        except DjangoValidationError:
                            row_errors.append(f"Invalid phone number: {phone_number}")

                    if row_errors:
                        errors.append({"row": idx, "values": row, "errors": row_errors})
                    else:
                        rows_to_create.append({
                            "company_name": company_name,
                            "email": email,
                            "address": address,
                            "phone_number": phone_number,
                        })

                if errors:
                    return Response({
                        "message": "Errors found in the uploaded file. No vendors were created or updated.",
                        "errors": errors
                    }, status=400)

                vendors_created = 0
                vendors_updated = 0
                create_instances = []

                try:
                    with transaction.atomic():
                        # Fetch all existing vendors with matching company_name and email
                        company_names = [row["company_name"] for row in rows_to_create]
                        emails = [row["email"] for row in rows_to_create]
                        existing_vendors = Vendor.objects.filter(company_name__in=company_names, email__in=emails)
                        existing_lookup = {(v.company_name.lower(), v.email.lower()): v for v in existing_vendors}

                        for row in rows_to_create:
                            key = (row["company_name"].lower(), row["email"].lower())
                            if key in existing_lookup:
                                vendor = existing_lookup[key]
                                vendor.address = row["address"]
                                vendor.phone_number = row["phone_number"]
                                update_instances.append(vendor)
                                vendors_updated += 1
                            else:
                                create_instances.append(Vendor(**row))
                                vendors_created += 1

                        # Bulk update in batches
                        for i in range(0, len(update_instances), BATCH_SIZE):
                            Vendor.objects.bulk_update(update_instances[i:i+BATCH_SIZE], ["address", "phone_number"])
                        # Bulk create in batches
                        for i in range(0, len(create_instances), BATCH_SIZE):
                            Vendor.objects.bulk_create(create_instances[i:i+BATCH_SIZE])

                except Exception as e:
                    return Response({"error": f"Error processing Excel file: {str(e)}"}, status=400)

                return Response({
                    "message": f"Successfully created {vendors_created} vendors, updated {vendors_updated} vendors",
                    "errors": []
                }, status=201)

            except Exception as e:
                return Response({"error": f"Error processing Excel file: {str(e)}"}, status=400)
        else:
            return Response(serializer.errors, status=400)

    @action(detail=True, methods=['POST'])
    def upload_profile_picture(self, request, pk=None):
        vendor = self.get_object()
        profile_file = request.FILES.get("profile_picture")

        if not profile_file:
            return Response({"error": "No file provided"}, status=400)

        vendor.profile_picture = convert_to_base64(profile_file)
        vendor.save()

        return Response({
            "message": "Profile picture uploaded successfully",
            "profile_picture": vendor.profile_picture
        }, status=200)


@extend_schema_view(
    list=extend_schema(tags=['Products']),
    retrieve=extend_schema(tags=['Products']),
    create=extend_schema(tags=['Products']),
    update=extend_schema(tags=['Products']),
    partial_update=extend_schema(tags=['Products']),
    destroy=extend_schema(tags=['Products']),
)
class ProductViewSet(SearchDeleteViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    app_label = "purchase"
    model_name = "product"
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    search_fields = ['product_name', 'product_category', "unit_of_measure__unit_name", ]
    filterset_fields = ["unit_of_measure__unit_name",]
    action_permission_map = {
        **basic_action_permission_map,
        "upload_excel": "create",
        "delete_all_products": "delete",
        "download_template": "view",
    }

    @action(detail=False, methods=['POST'], serializer_class=ExcelUploadSerializer)
    def upload_excel(self, request):
        serializer = ExcelUploadSerializer(data=request.data)
        if serializer.is_valid():
            excel_file = serializer.validated_data['file']
            check_for_duplicates = serializer.validated_data.get('check_for_duplicates', False)

            if not excel_file or not isinstance(excel_file, InMemoryUploadedFile):
                return Response({"error": "No file provided or invalid file format"}, status=400)

            try:
                workbook = load_workbook(excel_file)
                if 'Products' not in workbook.sheetnames:
                    return Response({"error": "No 'Products' sheet found in the uploaded file."}, status=400)
                sheet = workbook['Products']

                # Check if the sheet is empty (only header or no data)
                if sheet.max_row < 2:
                    return Response({"error": "The 'Products' sheet is empty."}, status=400)

                valid_product_categories = [choice[0] for choice in PRODUCT_CATEGORY]
                errors = []
                rows_to_create = []
                update_instances = []

                for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    product_name, product_description, product_category, unit_of_measure_name = row[:4]
                    row_errors = []

                    # Check if the entire row is empty
                    if all(cell is None or str(cell).strip() == "" for cell in (product_name, product_description, product_category, unit_of_measure_name)):
                        row_errors.append("Entire row is empty.")
                        errors.append({"row": idx, "values": row, "errors": row_errors})
                        continue

                    # Check for empty columns
                    empty_columns = []
                    if not product_name or str(product_name).strip() == "":
                        empty_columns.append("product_name")
                    if not product_description or str(product_description).strip() == "":
                        empty_columns.append("product_description")
                    if not product_category or str(product_category).strip() == "":
                        empty_columns.append("product_category")
                    if not unit_of_measure_name or str(unit_of_measure_name).strip() == "":
                        empty_columns.append("unit_of_measure")
                    if empty_columns:
                        row_errors.append(f"Missing required fields: {', '.join(empty_columns)}. Row values: {row}")

                    product_category_slug = slugify(product_category) if product_category else ""

                    if product_category and product_category_slug not in valid_product_categories:
                        row_errors.append(
                            f"Invalid category '{product_category}' for {product_name}. "
                            f"Valid categories are: {(', '.join(valid_product_categories)).title().replace('-', ' ')}."
                        )

                    try:
                        unit_of_measure = UnitOfMeasure.objects.get(unit_name=unit_of_measure_name)
                        unit_of_measure_id = unit_of_measure.id
                    except UnitOfMeasure.DoesNotExist:
                        row_errors.append(f"Unit of measure '{unit_of_measure_name}' does not exist for {product_name}.")
                        unit_of_measure_id = None

                    if row_errors:
                        errors.append({"row": idx, "errors": row_errors})
                    else:
                        rows_to_create.append({
                            "product_name": product_name,
                            "product_description": product_description,
                            "product_category": product_category_slug,
                            "unit_of_measure_id": unit_of_measure_id
                        })

                if errors:
                    return Response({
                        "message": "Errors found in the uploaded file. No products were created or updated.",
                        "errors": errors
                    }, status=400)

                products_created = 0
                products_updated = 0
                create_instances = []
                BATCH_SIZE = 500

                try:
                    with transaction.atomic():
                        if check_for_duplicates:
                            product_names = [row["product_name"] for row in rows_to_create]
                            product_categories = [row["product_category"] for row in rows_to_create]
                            existing_products = Product.objects.filter(
                                product_name__in=product_names,
                                product_category__in=product_categories
                            )
                            existing_lookup = {
                                (p.product_name.lower(), p.product_category.lower()): p for p in existing_products
                            }

                            for row in rows_to_create:
                                key = (row["product_name"].lower(), row["product_category"].lower())
                                if key in existing_lookup:
                                    product = existing_lookup[key]
                                    product.product_description = row["product_description"]
                                    product.unit_of_measure_id = row["unit_of_measure_id"]
                                    update_instances.append(product)
                                    products_updated += 1
                                else:
                                    create_instances.append(Product(**row))
                                    products_created += 1

                            # Bulk update in batches
                            for i in range(0, len(update_instances), BATCH_SIZE):
                                Product.objects.bulk_update(
                                    update_instances[i:i+BATCH_SIZE],
                                    ["product_description", "unit_of_measure_id"]
                                )
                            # Bulk create in batches
                            for i in range(0, len(create_instances), BATCH_SIZE):
                                Product.objects.bulk_create(create_instances[i:i+BATCH_SIZE])
                        else:
                            # Only create new products in batches
                            create_instances = [Product(**row) for row in rows_to_create]
                            for i in range(0, len(create_instances), BATCH_SIZE):
                                Product.objects.bulk_create(create_instances[i:i+BATCH_SIZE])
                            products_created = len(create_instances)
                except Exception as e:
                    return Response({"error": f"Error processing Excel file: {str(e)}"}, status=400)

                return Response({
                    "message": f"Successfully created {products_created} products, updated {products_updated} products",
                    "errors": errors
                }, status=201 if not errors else 400)

            except Exception as e:
                return Response({"error": f"Error processing Excel file: {str(e)}"}, status=400)
        else:
            return Response(serializer.errors, status=400)

    @action(detail=False, methods=['GET'], url_path='download-template')
    def download_template(self, request):
        """
        Endpoint to download a template Excel file for product import.
        The workbook will have:
        - An 'Instructions' sheet as the first sheet.
        - A 'Products' sheet for data entry (and import).
        - A list of all available unit_of_measure names at the time of download.
        - Data validation for each column.
        """
        wb = Workbook()
        ws_instructions = wb.active
        ws_instructions.title = "Instructions"
        ws_instructions["A1"] = "Instructions for Filling the Products Sheet:"
        ws_instructions["A2"] = "1. Fill each row in the 'Products' sheet with product details."
        ws_instructions["A3"] = (
            "2. 'product_category' should match one of the allowed categories: "
            + ", ".join([choice[0] for choice in PRODUCT_CATEGORY])
        )
        ws_instructions["A4"] = "3. 'unit_of_measure' should match an existing unit name (see below)."
        ws_instructions["A5"] = "4. Do not modify the header row."
        ws_instructions["A7"] = "Available unit_of_measure names:"

        # Add all unit_of_measure names starting from A8
        unit_names = list(UnitOfMeasure.objects.values_list("unit_name", flat=True))
        for idx, name in enumerate(unit_names, start=8):
            ws_instructions[f"A{idx}"] = name

        # Add the Products sheet as the second sheet
        ws_products = wb.create_sheet(title="Products")
        headers = [
            "product_name",
            "product_description",
            "product_category",
            "unit_of_measure",
        ]
        ws_products.append(headers)

        # Data validation for product_name (required, not blank)
        name_dv = DataValidation(
            type="custom",
            formula1='=LEN(TRIM(A2))>0',
            showErrorMessage=True,
            errorTitle="Required Field",
            error="Product name is required."
        )
        ws_products.add_data_validation(name_dv)
        name_dv.add("A2:A1048576")

        # Data validation for product_description (required, not blank)
        desc_dv = DataValidation(
            type="custom",
            formula1='=LEN(TRIM(B2))>0',
            showErrorMessage=True,
            errorTitle="Required Field",
            error="Product description is required."
        )
        ws_products.add_data_validation(desc_dv)
        desc_dv.add("B2:B1048576")

        # Data validation for product_category (dropdown, type-able, must match slugified value)
        category_slugs = [slugify(choice[0]) for choice in PRODUCT_CATEGORY]
        # Create a hidden sheet to store the list for dropdown
        ws_hidden = wb.create_sheet(title="ValidationLists")
        for idx, slug in enumerate(category_slugs, start=1):
            ws_hidden[f"A{idx}"] = slug
        ws_hidden.sheet_state = 'hidden'
        # Reference for dropdown
        category_range = f"ValidationLists!$A$1:$A${len(category_slugs)}"
        cat_dv = DataValidation(
            type="list",
            formula1=f"={category_range}",
            allow_blank=False,
            showDropDown=True,
            showErrorMessage=True,
            errorTitle="Invalid Category",
            error="Category must match one of the allowed slug values."
        )
        ws_products.add_data_validation(cat_dv)
        cat_dv.add("C2:C1048576")

        # Data validation for unit_of_measure (dropdown, must match existing unit name)
        for idx, name in enumerate(unit_names, start=1):
            ws_hidden[f"B{idx}"] = name
        unit_range = f"ValidationLists!$B$1:$B${len(unit_names)}"
        unit_dv = DataValidation(
            type="list",
            formula1=f"={unit_range}",
            allow_blank=False,
            showDropDown=True,
            showErrorMessage=True,
            errorTitle="Invalid Unit",
            error="Unit of measure must match one of the available units."
        )
        ws_products.add_data_validation(unit_dv)
        unit_dv.add("D2:D1048576")

        # Save workbook to a BytesIO stream
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=product_import_template.xlsx'
        return response

    @action(detail=False, methods=['DELETE'], permission_classes=[IsAdminUser], url_path='delete-all',
            url_name='delete_all_products')
    def delete_all_products(self, request):
        deleted_count, _ = Product.objects.all().delete()

        return Response(
            {"message": f"Successfully deleted {deleted_count} products."},
            status=status.HTTP_200_OK
        )


@extend_schema_view(
    list=extend_schema(tags=['Purchase Requests']),
    retrieve=extend_schema(tags=['Purchase Requests']),
    create=extend_schema(tags=['Purchase Requests']),
    update=extend_schema(tags=['Purchase Requests']),
    partial_update=extend_schema(tags=['Purchase Requests']),
    destroy=extend_schema(tags=['Purchase Requests']),
)
class PurchaseRequestViewSet(SearchDeleteViewSet):
    queryset = PurchaseRequest.objects.all()
    serializer_class = PurchaseRequestSerializer
    # Required by permission class
    app_label = "purchase"
    model_name = "purchaserequest"
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    search_fields = ['id', 'status', "vendor__company_name", "currency__currency_name",
                     "requesting_location__location_name"]
    filterset_fields = ['status', "requesting_location__id", "requester__user_id", 'date_created']
    # Map DRF actions to your permission names
    action_permission_map = {
        **basic_action_permission_map,
        "draft_list": "view",
        "pending_list": "view",
        "approved_list": "view",
        "rejected_list": "view",
        "convert_to_rfq": "create",
        "submit": "edit",
        "approve": "approve",
        "reject": "reject",
    }


    def perform_create(self, serializer):
        # Ensure the user is a TenantUser
        user = self.request.user
        # Ensure we are operating within the correct tenant schema
        tenant = self.request.tenant  # Get the current tenant

        with tenant_context(tenant):  # Switch to the tenant's schema
            try:
                tenant_user = TenantUser.objects.get(user_id=user.id)
            except ObjectDoesNotExist:
                raise serializers.ValidationError("Requester must be a TenantUser within the tenant schema.")

        serializer.save(requester=tenant_user)

    def perform_update(self, serializer):
        # Ensure the user is a TenantUser
        user = self.request.user
        # Ensure we are operating within the correct tenant schema
        tenant = self.request.tenant
        with tenant_context(tenant):  # Switch to the tenant's schema
            try:
                tenant_user = TenantUser.objects.get(user_id=user.id)
            except ObjectDoesNotExist:
                raise serializers.ValidationError("Requester must be a TenantUser within the tenant schema.")
        serializer.save(requester=tenant_user)

    @action(detail=False, methods=['get'])
    def draft_list(self, request):
        queryset = PurchaseRequest.pr_draft.all()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def pending_list(self, request):
        queryset = PurchaseRequest.pr_pending.all()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def approved_list(self, request):
        queryset = PurchaseRequest.pr_approved.all()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def rejected_list(self, request):
        queryset = PurchaseRequest.pr_rejected.all()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['put', 'patch'])
    def submit(self, request, pk=None):
        purchase_request = self.get_object()
        purchase_request.submit()
        return Response({'status': 'submitted'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['put', 'patch'])
    def approve(self, request, pk=None):
        purchase_request = self.get_object()
        # if request.user.has_perm('approve_purchase_request'):
        purchase_request.approve()
        return Response({'status': 'approved'}, status=status.HTTP_200_OK)
        # return Response({'status': 'permission denied'}, status=status.HTTP_403_FORBIDDEN)

    @action(detail=True, methods=['put', 'patch'])
    def reject(self, request, pk=None):
        purchase_request = self.get_object()
        # if request.user.has_perm('reject_purchase_request'):
        purchase_request.reject()
        return Response({'status': 'rejected'}, status=status.HTTP_200_OK)
        # return Response({'status': 'permission denied'}, status=status.HTTP_403_FORBIDDEN)

    @action(detail=True, methods=['post'])
    def convert_to_rfq(self, request, pk=None):
        try:
            # Get the approved purchase request
            purchase_request = self.get_object()

            if purchase_request.status != 'approved':
                return Response({"detail": "Only approved purchase requests can be converted to RFQs."},
                                status=status.HTTP_400_BAD_REQUEST)

            # Create the RFQ
            rfq = RequestForQuotation.objects.create(
                purchase_request=purchase_request,
                vendor=purchase_request.vendor,  # Assuming vendor is used
                expiry_date=None,  # Can be set dynamically or left blank
                currency=purchase_request.currency,  # Can be set dynamically or left blank
                status='draft',
                date_created=timezone.now(),
                date_updated=timezone.now(),
                is_hidden=False,
            )

            # Create RFQ items from the PurchaseRequest items
            for pr_item in purchase_request.items.all():
                RequestForQuotationItem.objects.create(
                    request_for_quotation=rfq,
                    product=pr_item.product,
                    description=pr_item.description,
                    unit_of_measure=pr_item.unit_of_measure,
                    qty=pr_item.qty,
                    estimated_unit_price=pr_item.estimated_unit_price,
                )

            return Response({
                "detail": "RFQ created successfully",
                "rfq_id": rfq.id
            }, status=status.HTTP_201_CREATED)

        except PurchaseRequest.DoesNotExist:
            return Response({"detail": "Purchase request not found."}, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({"detail": f"An error occurred: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

@extend_schema_view(
    list=extend_schema(tags=['Purchase Request Items']),
    retrieve=extend_schema(tags=['Purchase Request Items']),
    create=extend_schema(tags=['Purchase Request Items']),
    update=extend_schema(tags=['Purchase Request Items']),
    partial_update=extend_schema(tags=['Purchase Request Items']),
    destroy=extend_schema(tags=['Purchase Request Items']),
)
class PurchaseRequestItemViewSet(SearchViewSet):
    queryset = PurchaseRequestItem.objects.all()
    serializer_class = PurchaseRequestItemSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['purchase_request__id']
    search_fields = ["product__product_name", "unit_of_measure__unit_name", "purchase_request__id"]

@extend_schema_view(
    list=extend_schema(tags=['Request For Quotation']),
    retrieve=extend_schema(tags=['Request For Quotation']),
    create=extend_schema(tags=['Request For Quotation']),
    update=extend_schema(tags=['Request For Quotation']),
    partial_update=extend_schema(tags=['Request For Quotation']),
    destroy=extend_schema(tags=['Request For Quotation']),
)
class RequestForQuotationViewSet(SearchDeleteViewSet):
    queryset = RequestForQuotation.objects.all()
    serializer_class = RequestForQuotationSerializer
    app_label = "purchase"
    model_name = "requestforquotation"
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    filterset_fields = ['status', "purchase_request__id", 'date_created']
    search_fields = ["vendor__company_name", 'status', "purchase_request__id"]
    action_permission_map = {
        **basic_action_permission_map,
        "check_rfq_editable": "view",
        "check_rfq_mailable": "view",
        "submit": "edit",
        "approve": "approve",
        "reject": "reject",
        "draft_list": "view",
        "toggle_hidden_status": "edit",
        "pending_list": "view",
        "approved_list": "view",
        "rejected_list": "view",
        "convert_to_po": "create",
    }


    @action(detail=True, methods=['get'])
    def check_rfq_editable(self, rfq):
        """Check if the RFQ is editable (not submitted or rejected)."""
        if rfq.is_submitted:
            return False, 'This RFQ has already been submitted and cannot be edited.'
        return True, ''

    @action(detail=True, methods=['get'])
    def check_rfq_mailable(self, rfq):
        """Check if the RFQ meets the criteria to be sent to vendors (not draft or rejected)."""
        if rfq.status in ['rejected', 'draft']:
            return False, 'This RFQ cannot be sent as it has been rejected or not submitted.'
        if rfq.is_expired:
            return False, 'This RFQ has expired and cannot be sent.'
        return True, ''

    # for sending RFQs to vendor emails
    @action(detail=True, methods=['post'])
    def send_email(self, request, pk=None):
        rfq = self.get_object()

        # Check if the RFQ has expired
        if rfq.is_expired:
            return Response({'error': 'Cannot send expired RFQ.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Generate the PDF
            pdf_response = generate_model_pdf(rfq)
            if not pdf_response:
                return Response({'error': 'Error generating PDF.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            pdf_content = pdf_response.content

            # Create and send the email
            subject = f"Request for Quotation: {rfq.id}"
            body = (f"Please find attached the RFQ {rfq.id}. The deadline"
                    f" for response is {rfq.expiry_date.strftime('%Y-%m-%d') if rfq.expiry_date else 'None'}.")
            # email = EmailMessage(subject, body, settings.EMAIL_HOST_USER, [rfq.vendor.email])
            # email.attach(f"RFQ_{rfq.id}.pdf", pdf_content, 'application/pdf')

            mailto_link = f'mailto:{rfq.vendor.email}?subject={quote(subject)}&body={quote(body)}'

            # Prepare the response
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="RFQ_{rfq.id}.pdf"'

            # Embed the mailto link in the response headers for the front-end to use
            response['X-Mailto-Link'] = mailto_link

            return response

            # email.send()

            # return Response({'status': 'email sent'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['put', 'patch'])
    def submit(self, request, pk=None):
        rfq = self.get_object()
        editable, message = self.check_rfq_editable(rfq)

        if not editable:
            return Response({'error': message}, status=status.HTTP_400_BAD_REQUEST)

        rfq.submit()
        return Response({'status': 'pending'})

    @action(detail=True, methods=['put', 'patch'])
    def approve(self, request, pk=None):
        rfq = self.get_object()
        if request.user.has_perm('approve_request_for_quotation'):
            rfq.approve()
            return Response({'status': 'approved'})
        return Response({'status': 'permission denied'}, status=403)

    @action(detail=True, methods=['put', 'patch'])
    def reject(self, request, pk=None):
        rfq = self.get_object()
        if request.user.has_perm('reject_request_for_quotation'):
            rfq.reject()
            return Response({'status': 'rejected'})
        return Response({'status': 'permission denied'}, status=403)

    @action(detail=False, methods=['get'])
    def draft_list(self, request):
        queryset = RequestForQuotation.rfq_draft.all()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def pending_list(self, request):
        queryset = RequestForQuotation.rfq_pending.all()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def approved_list(self, request):
        queryset = RequestForQuotation.rfq_approved.all()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def rejected_list(self, request):
        queryset = RequestForQuotation.rfq_rejected.all()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def get_queryset(self):
        queryset = super().get_queryset()  # Use the superclass queryset
        rfq_status = self.request.query_params.get('status')
        expired = self.request.query_params.get('expired')

        # Filter by status if provided
        if rfq_status:
            queryset = queryset.filter(status=rfq_status)

        # Filter by expiration status if provided
        if expired is not None:
            queryset = queryset.filter(expiry_date__lt=timezone.now()) if expired == 'true' else queryset.filter(
                expiry_date__gte=timezone.now())

        return queryset

    @action(detail=True, methods=['post'])
    def convert_to_po(self, request, pk=None):
        try:
            # Get the approved purchase request
            rfq = self.get_object()

            if rfq.status != 'approved':
                return Response({"detail": "Only approved Requests For Quotation can be converted to Purchase Orders."},
                                status=status.HTTP_400_BAD_REQUEST)
            if not rfq.actual_price:
                return Response({'error': 'Actual price is required to convert to PO.'},
                                status=status.HTTP_400_BAD_REQUEST)

            # Create the Purchase Order
            po = PurchaseOrder.objects.create(
                vendor=rfq.vendor,
                status='draft',
                created_by=request.user,
                currency=rfq.currency,
                is_hidden=False,
            )

            # Create po items from the RFQ items
            for rfq_item in rfq.items.all():
                PurchaseOrderItem.objects.create(
                    purchase_order=po,
                    product=rfq_item.product,
                    description=rfq_item.description,
                    qty=rfq_item.qty,
                    unit_of_measure=rfq_item.unit_of_measure,
                    estimated_unit_price=rfq_item.estimated_unit_price,
                )

            return Response({
                "detail": "Purchase Order created successfully",
                "po_id": po.id
            }, status=status.HTTP_201_CREATED)

        except RequestForQuotation.DoesNotExist:
            return Response({"detail": "Request For Quotation not found."}, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({"detail": f"An error occurred: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


@extend_schema_view(
    list=extend_schema(tags=['Request For Quotation Items']),
    retrieve=extend_schema(tags=['Request For Quotation Items']),
    create=extend_schema(tags=['Request For Quotation Items']),
    update=extend_schema(tags=['Request For Quotation Items']),
    partial_update=extend_schema(tags=['Request For Quotation Items']),
    destroy=extend_schema(tags=['Request For Quotation Items']),
)
class RequestForQuotationItemViewSet(SearchViewSet):
    queryset = RequestForQuotationItem.objects.all()
    serializer_class = RequestForQuotationItemSerializer
    permission_classes = [permissions.IsAuthenticated]
    search_fields = ["product__product_name", "unit_of_measure__unit_name",]
    filterset_fields = ["product__product_name", "unit_of_measure__unit_name", "request_for_quotation__id"]


@extend_schema_view(
    list=extend_schema(tags=['Purchase Orders']),
    retrieve=extend_schema(tags=['Purchase Orders']),
    create=extend_schema(tags=['Purchase Orders']),
    update=extend_schema(tags=['Purchase Orders']),
    partial_update=extend_schema(tags=['Purchase Orders']),
    destroy=extend_schema(tags=['Purchase Orders']),
)
class PurchaseOrderViewSet(SearchDeleteViewSet):
    queryset = PurchaseOrder.objects.all()
    serializer_class = PurchaseOrderSerializer
    app_label = "purchase"
    model_name = "purchaseorder"
    permission_classes = [permissions.IsAuthenticated, HasModulePermission]
    filterset_fields = ['status', "destination_location__id", "created_by__user_id", 'date_created']
    search_fields = ['status', "vendor__company_name", "related_rfq__id", "destination_location__location_name"]
    lookup_field = 'id'
    lookup_url_kwarg = 'id'
    action_permission_map = {
        **basic_action_permission_map,
        "convert_to_incoming_product": "create",
        "submit": "edit",
        "complete": "edit",
        "cancel": "edit",
        "draft_list": "view",
        "awaiting_list": "view",
        "cancelled_list": "view",
        "completed_list": "view",
        "get_unrelated_po": "view"
    }

    def perform_create(self, serializer):
        # Ensure the user is a TenantUser
        user = self.request.user
        # Ensure we are operating within the correct tenant schema
        tenant = self.request.tenant  # Get the current tenant

        with tenant_context(tenant):  # Switch to the tenant's schema
            try:
                tenant_user = TenantUser.objects.get(user_id=user.id)
            except ObjectDoesNotExist:
                raise serializers.ValidationError("This user must be a TenantUser within the tenant schema.")

        serializer.save(created_by=tenant_user)


    @action(detail=True, methods=['get'])
    def check_po_editable(self, po):
        """Check if the PO is editable (not submitted or rejected)."""
        if po.is_submitted:
            return False, 'This purchase order has already been submitted and cannot be edited.'
        return True, ''

    @action(detail=True, methods=['get'])
    def check_po_mailable(self, po):
        """Check if the PO meets the criteria to be sent to vendors (not draft or rejected)."""
        if po.status in ['cancelled', 'draft']:
            return False, 'This purchase order cannot be sent as it has been rejected or not submitted.'
        if po.is_expired:
            return False, 'This purchase order has expired and cannot be sent.'
        return True, ''

    # for sending POs to vendor emails
    @action(detail=True, methods=['post'])
    def send_email(self, request, pk=None):
        po = self.get_object()

        try:
            # Generate the PDF
            pdf_response = generate_model_pdf(po)
            if not pdf_response:
                return Response({'error': 'Error generating PDF.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            pdf_content = pdf_response.content

            # Create and send the email
            subject = f"Purchase Order: {po.id}"
            body = f"Please find attached the RFQ {po.id}."
            # email = EmailMessage(subject, body, settings.EMAIL_HOST_USER, [po.vendor.email])
            # email.attach(f"PO_{po.id}.pdf", pdf_content, 'application/pdf')

            mailto_link = f'mailto:{po.vendor.email}?subject={quote(subject)}&body={quote(body)}'

            # Prepare the response
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="PO_{po.id}.pdf"'

            # Embed the mailto link in the response headers for the front-end to use
            response['X-Mailto-Link'] = mailto_link

            return response

            # email.send()

            # return Response({'status': 'email sent'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(methods=['get'], detail=False)
    def get_unrelated_po(self, request):
        unrelated_po = PurchaseOrder.objects.filter(incoming_product__isnull=True, status="completed").distinct()
        serializer = self.get_serializer(unrelated_po, many=True)
        return Response(serializer.data)


    @action(detail=True, methods=['post'])
    def convert_to_incoming_product(self, request, pk=None):
        try:
            po = self.get_object()

            if po.status != 'completed':
                return Response({"detail": "Only approved Requests For Quotation can be converted to Purchase Orders."},
                                status=status.HTTP_400_BAD_REQUEST)
            if not po.actual_price:
                return Response({'error': 'Actual price is required to convert to PO.'},
                                status=status.HTTP_400_BAD_REQUEST)

            # Create the Purchase Order
            incoming_product = IncomingProduct.objects.create(
                receipt_type="vendor_receipt",
                related_po=po,
                supplier=po.vendor,
                source_location="SUPP00001",
                destination_location=Location.get_active_locations(),
                status='draft'
            )

            # Create po items from the RFQ items
            for po_item in po.items.all():
                IncomingProductItem.objects.create(
                    incoming_product=incoming_product,
                    product=po_item.product,
                    expected_quantity=po_item.qty
                )

            return Response({
                "detail": "Incoming product created successfully",
                "incoming_product_id": po.id
            }, status=status.HTTP_201_CREATED)

        except PurchaseOrder.DoesNotExist:
            return Response({"detail": "Purchase Order not found."}, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({"detail": f"An error occurred: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


    @action(detail=True, methods=['put', 'patch'])
    def submit(self, request, pk=None):
        po = self.get_object()
        editable, message = self.check_po_editable(po)

        if not editable:
            return Response({'error': message}, status=status.HTTP_400_BAD_REQUEST)

        po.submit()
        return Response({'status': 'awaiting'})

    @action(detail=True, methods=['put', 'patch'])
    def complete(self, request, pk=None):
        po = self.get_object()
        if request.user.has_perm('complete_purchase_order'):
            po.complete()
            return Response({'status': 'completed'})
        return Response({'status': 'permission denied'}, status=403)

    @action(detail=True, methods=['put', 'patch'])
    def cancel(self, request, pk=None):
        po = self.get_object()
        if request.user.has_perm('cancel_purchase_order'):
            po.cancel()
            return Response({'status': 'cancelled'})
        return Response({'status': 'permission denied'}, status=403)


    @action(detail=False, methods=['get'])
    def draft_list(self, request):
        queryset = PurchaseOrder.po_draft.all()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def awaiting_list(self, request):
        queryset = PurchaseOrder.po_awaiting.all()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def cancelled_list(self, request):
        queryset = PurchaseOrder.po_cancelled.all()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def completed_list(self, request):
        queryset = PurchaseOrder.po_completed.all()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


@extend_schema_view(
    list=extend_schema(tags=['Purchase Order Items']),
    retrieve=extend_schema(tags=['Purchase Order Items']),
    create=extend_schema(tags=['Purchase Order Items']),
    update=extend_schema(tags=['Purchase Order Items']),
    partial_update=extend_schema(tags=['Purchase Order Items']),
    destroy=extend_schema(tags=['Purchase Order Items']),
)
class PurchaseOrderItemViewSet(SearchViewSet):
    queryset = PurchaseOrderItem.objects.all()
    serializer_class = PurchaseOrderItemSerializer
    permission_classes = [permissions.IsAuthenticated]
    search_fields = ["product__product_name", "unit_of_measure__unit_name",]
    filterset_fields = ["product__product_name", "unit_of_measure__unit_name", "purchase_order__id"]
